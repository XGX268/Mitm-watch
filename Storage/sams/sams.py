cat > sams.py <<'PY'
import csv, hashlib, re, subprocess, time
from pathlib import Path
from xml.etree import ElementTree as ET
from PIL import Image
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage

PACKAGE = "cn.samsclub.app"
MAX_SCROLLS = 500
NO_NEW_LIMIT = 8
WAIT_AFTER_SCROLL = 1.2

PRICE_RE = re.compile(r'^[¥￥]\s*[\d,]+(?:\.\d{1,2})?$')
BOUNDS_RE = re.compile(r'\[(\d+),(\d+)\]\[(\d+),(\d+)\]')

STOP = {
    "首页","首頁","分類","分类","发现","發現","购物车","購物車","我的",
    "销量","銷量","价格","價格","全部","展开","展開",
    "满赠","滿贈","礼盒","禮盒","极速达","極速達","全国配","全國配"
}

def run(cmd, check=True, text=True):
    return subprocess.run(
        cmd,
        shell=True,
        check=check,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=text
    )

def adb(args, check=True, text=True):
    return run("adb " + args, check, text)

def parse_bounds(s):
    m = BOUNDS_RE.fullmatch(s or "")
    return tuple(map(int, m.groups())) if m else None

def cy(b):
    return (b[1] + b[3]) / 2

def area(b):
    return max(0, b[2]-b[0]) * max(0, b[3]-b[1])

def clean(t):
    return re.sub(r"\s+", " ", (t or "").strip())

def is_price(t):
    return bool(PRICE_RE.match(clean(t)))

def looks_name(t):
    t = clean(t)

    if not (4 <= len(t) <= 100):
        return False

    if t in STOP:
        return False

    if is_price(t):
        return False

    if not re.search(r"[\u4e00-\u9fffA-Za-z]", t):
        return False

    return True

def parse_nodes(xml):
    root = ET.fromstring(xml)

    out = []

    for n in root.iter("node"):
        b = parse_bounds(n.attrib.get("bounds", ""))

        if not b:
            continue

        t = clean(
            n.attrib.get("text", "")
            or n.attrib.get("content-desc", "")
        )

        out.append({
            "text": t,
            "bounds": b,
            "class": n.attrib.get("class", "")
        })

    return out

def title_score(t, gap):
    score = abs(gap - 120)

    if re.search(r"[；;。！？!?]", t):
        score += 100

    if "，" in t or "," in t:
        score += 45

    if len(t) > 45:
        score += 45

    if re.search(
        r"\d+(?:\.\d+)?\s*(?:g|kg|ml|mL|L|升|克|千克|斤|入|枚|件|瓶|罐|包|袋|片|粒|支|盒)",
        t,
        re.I
    ):
        score -= 35

    return score

def find_products(nodes, screen_w):
    prices = [n for n in nodes if is_price(n["text"])]
    names = [n for n in nodes if looks_name(n["text"])]

    out = []

    for p in prices:
        py = cy(p["bounds"])

        candidates = []

        for n in names:
            gap = py - cy(n["bounds"])

            if not (20 <= gap <= 360):
                continue

            if abs(n["bounds"][0] - p["bounds"][0]) > screen_w * 0.28:
                continue

            candidates.append(
                (title_score(n["text"], gap), n)
            )

        if not candidates:
            continue

        candidates.sort(key=lambda x: x[0])
        name = candidates[0][1]

        images = []

        for n in nodes:
            if "image" not in n["class"].lower():
                continue

            b = n["bounds"]

            if b[2] > name["bounds"][0] + 80:
                continue

            target_y = (cy(name["bounds"]) + py) / 2

            if abs(cy(b) - target_y) > 320:
                continue

            if b[2]-b[0] < 70 or b[3]-b[1] < 70:
                continue

            score = (
                abs(cy(b) - target_y)
                - min(area(b), 250000) / 8000
            )

            images.append((score, n))

        images.sort(key=lambda x: x[0])

        out.append({
            "name": name["text"],
            "price": clean(p["text"]).replace(" ", ""),
            "name_bounds": name["bounds"],
            "price_bounds": p["bounds"],
            "image_bounds": images[0][1]["bounds"] if images else None,
            "y": py
        })

    seen = set()
    final = []

    for x in sorted(out, key=lambda z: z["y"]):
        k = (x["name"], x["price"])

        if k in seen:
            continue

        seen.add(k)
        final.append(x)

    return final

def fallback_image_bounds(prod, sw, sh):
    nb = prod["name_bounds"]
    pb = prod["price_bounds"]

    top = max(0, min(nb[1], pb[1]) - 150)
    bottom = min(sh, max(nb[3], pb[3]) + 120)

    right = max(10, nb[0] - 20)
    width = min(int(sw * 0.28), right)
    left = max(0, right - width)

    return (left, top, right, bottom)

def make_key(name, price):
    return hashlib.sha1(
        (name + "|" + price).encode()
    ).hexdigest()[:16]

def get_outdir():
    downloads = Path.home() / "storage" / "downloads"

    if downloads.exists():
        p = downloads / "sams_export"
    else:
        p = Path.home() / "sams_export"

    (p / "images").mkdir(parents=True, exist_ok=True)
    (p / "debug").mkdir(exist_ok=True)

    return p

def save_csv(rows, p):
    with (p / "products.csv").open(
        "w",
        newline="",
        encoding="utf-8-sig"
    ) as f:

        w = csv.writer(f)

        w.writerow([
            "商品圖",
            "商品名稱",
            "價格"
        ])

        for r in rows:
            w.writerow([
                str(r["image"]),
                r["name"],
                r["price"]
            ])

def save_excel(rows, p):
    wb = Workbook()
    ws = wb.active

    ws.title = "商品"

    ws.append([
        "商品圖",
        "商品名稱",
        "價格"
    ])

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].width = 16

    for i, r in enumerate(rows, 2):

        ws.row_dimensions[i].height = 95

        ws.cell(i, 2, r["name"])
        ws.cell(i, 3, r["price"])

        try:
            img = XLImage(str(r["image"]))
            img.width = 95
            img.height = 95

            ws.add_image(img, f"A{i}")
        except Exception:
            pass

    path = p / "products.xlsx"

    wb.save(path)

    return path

def main():
    devices = adb(
        "devices",
        check=False
    ).stdout

    if "\tdevice" not in devices:
        print("ADB 未連線")
        print("先確認 adb devices 能看到 device")
        return

    wm = adb(
        "shell wm size"
    ).stdout

    m = re.search(
        r"Physical size:\s*(\d+)x(\d+)",
        wm
    )

    if not m:
        print("無法取得螢幕解析度")
        return

    sw, sh = map(int, m.groups())

    p = get_outdir()

    print("螢幕:", sw, "x", sh)
    print("開啟山姆 App")

    adb(
        f"shell monkey -p {PACKAGE} 1",
        check=False
    )

    time.sleep(2)

    rows = {}
    no_new = 0

    for page in range(MAX_SCROLLS):

        adb(
            "shell uiautomator dump /sdcard/sams.xml",
            check=False
        )

        xml = adb(
            "shell cat /sdcard/sams.xml"
        ).stdout

        if f'package="{PACKAGE}"' not in xml:

            print("焦點不在山姆，重新切回")

            adb(
                f"shell monkey -p {PACKAGE} 1",
                check=False
            )

            time.sleep(1.2)

            adb(
                "shell uiautomator dump /sdcard/sams.xml",
                check=False
            )

            xml = adb(
                "shell cat /sdcard/sams.xml"
            ).stdout

        shot = (
            p
            / "debug"
            / f"page_{page:04d}.png"
        )

        with shot.open("wb") as f:

            subprocess.run(
                "adb exec-out screencap -p",
                shell=True,
                stdout=f,
                stderr=subprocess.PIPE
            )

        screen = Image.open(
            shot
        ).convert("RGB")

        nodes = parse_nodes(xml)

        products = find_products(
            nodes,
            sw
        )

        new_count = 0

        for prod in products:

            k = make_key(
                prod["name"],
                prod["price"]
            )

            if k in rows:
                continue

            b = (
                prod["image_bounds"]
                or fallback_image_bounds(
                    prod,
                    sw,
                    sh
                )
            )

            x1, y1, x2, y2 = b

            b = (
                max(0, x1-6),
                max(0, y1-6),
                min(sw, x2+6),
                min(sh, y2+6)
            )

            image_path = (
                p
                / "images"
                / f"{k}.jpg"
            )

            crop = screen.crop(b)

            if (
                crop.width < 60
                or crop.height < 60
            ):
                crop = screen.crop(
                    fallback_image_bounds(
                        prod,
                        sw,
                        sh
                    )
                )

            crop.save(
                image_path,
                "JPEG",
                quality=90
            )

            rows[k] = {
                "name": prod["name"],
                "price": prod["price"],
                "image": image_path
            }

            new_count += 1

            print(
                "+",
                prod["name"],
                prod["price"]
            )

        save_csv(
            list(rows.values()),
            p
        )

        print(
            "page:",
            page,
            "detected:",
            len(products),
            "new:",
            new_count,
            "total:",
            len(rows)
        )

        if new_count == 0:
            no_new += 1
        else:
            no_new = 0

        if no_new >= NO_NEW_LIMIT:
            print("連續多屏沒有新商品，停止")
            break

        adb(
            f"shell input swipe "
            f"{int(sw*0.5)} "
            f"{int(sh*0.78)} "
            f"{int(sw*0.5)} "
            f"{int(sh*0.30)} "
            f"550",
            check=False
        )

        time.sleep(
            WAIT_AFTER_SCROLL
        )

    xlsx = save_excel(
        list(rows.values()),
        p
    )

    print("")
    print("完成")
    print("商品數:", len(rows))
    print("Excel:", xlsx)
    print("CSV:", p / "products.csv")
    print("圖片:", p / "images")

if __name__ == "__main__":
    main()
PY