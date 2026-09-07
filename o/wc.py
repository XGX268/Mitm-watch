import csv
import hashlib
import re
import subprocess
import time
from pathlib import Path
from xml.etree import ElementTree as ET

from PIL import Image
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage


WECHAT_PACKAGE = "com.tencent.mm"

MAX_SCROLLS = 400
NO_NEW_LIMIT = 10
WAIT_AFTER_SCROLL = 1.5

PRICE_RE = re.compile(
    r"(?:¥|￥)\s*[\d,]+(?:\.\d{1,2})?"
)

BOUNDS_RE = re.compile(
    r"\[(\d+),(\d+)\]\[(\d+),(\d+)\]"
)

STOP = {
    "微信",
    "小程序",
    "首頁",
    "首页",
    "購物車",
    "购物车",
    "我的",
    "分類",
    "分类",
    "搜索",
    "搜尋",
    "客服",
    "分享",
    "返回",
}


def run(cmd, check=True, text=True):
    return subprocess.run(
        cmd,
        shell=True,
        check=check,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=text,
    )


def adb(args, check=True, text=True):
    return run(
        "adb " + args,
        check=check,
        text=text,
    )


def clean(text):
    return re.sub(
        r"\s+",
        " ",
        (text or "").strip(),
    )


def parse_bounds(value):
    m = BOUNDS_RE.fullmatch(value or "")
    if not m:
        return None

    return tuple(
        map(int, m.groups())
    )


def cy(bounds):
    return (
        bounds[1] + bounds[3]
    ) / 2


def area(bounds):
    return (
        max(0, bounds[2] - bounds[0])
        * max(0, bounds[3] - bounds[1])
    )


def parse_nodes(xml):
    root = ET.fromstring(xml)

    nodes = []

    for node in root.iter("node"):
        bounds = parse_bounds(
            node.attrib.get(
                "bounds",
                ""
            )
        )

        if not bounds:
            continue

        text = clean(
            node.attrib.get(
                "text",
                ""
            )
            or node.attrib.get(
                "content-desc",
                ""
            )
        )

        nodes.append({
            "text": text,
            "bounds": bounds,
            "class": node.attrib.get(
                "class",
                ""
            ),
            "resource_id": node.attrib.get(
                "resource-id",
                ""
            ),
        })

    return nodes


def looks_like_name(text):
    text = clean(text)

    if not (
        3 <= len(text) <= 120
    ):
        return False

    if text in STOP:
        return False

    if PRICE_RE.fullmatch(text):
        return False

    if not re.search(
        r"[\u4e00-\u9fffA-Za-z]",
        text,
    ):
        return False

    return True


def find_products(nodes, screen_w):
    prices = []

    for node in nodes:
        if PRICE_RE.search(
            node["text"]
        ):
            prices.append(node)

    names = [
        node
        for node in nodes
        if looks_like_name(
            node["text"]
        )
    ]

    products = []

    for price_node in prices:
        py = cy(
            price_node["bounds"]
        )

        candidates = []

        for name_node in names:
            ny = cy(
                name_node["bounds"]
            )

            gap = py - ny

            # 商品名通常在價格上面
            if not (
                15 <= gap <= 420
            ):
                continue

            # 不要跨太遠的欄位配對
            if abs(
                name_node["bounds"][0]
                - price_node["bounds"][0]
            ) > screen_w * 0.40:
                continue

            score = abs(
                gap - 100
            )

            # 過長的描述降低優先級
            if len(
                name_node["text"]
            ) > 60:
                score += 40

            candidates.append(
                (
                    score,
                    name_node,
                )
            )

        if not candidates:
            continue

        candidates.sort(
            key=lambda x: x[0]
        )

        name_node = (
            candidates[0][1]
        )

        price_match = (
            PRICE_RE.search(
                price_node["text"]
            )
        )

        if not price_match:
            continue

        products.append({
            "name":
                name_node["text"],

            "price":
                price_match.group(0)
                .replace(" ", ""),

            "name_bounds":
                name_node["bounds"],

            "price_bounds":
                price_node["bounds"],

            "y":
                py,
        })

    # 去除同一屏重複
    final = []
    seen = set()

    for product in sorted(
        products,
        key=lambda x: x["y"],
    ):
        key = (
            product["name"],
            product["price"],
        )

        if key in seen:
            continue

        seen.add(key)
        final.append(product)

    return final


def get_image_bounds(
    product,
    sw,
    sh,
):
    name_b = (
        product["name_bounds"]
    )

    price_b = (
        product["price_bounds"]
    )

    top = max(
        0,
        min(
            name_b[1],
            price_b[1],
        ) - 180,
    )

    bottom = min(
        sh,
        max(
            name_b[3],
            price_b[3],
        ) + 100,
    )

    # 假設商品圖在商品文字左側
    right = max(
        10,
        name_b[0] - 10,
    )

    width = min(
        int(sw * 0.32),
        right,
    )

    left = max(
        0,
        right - width,
    )

    return (
        left,
        top,
        right,
        bottom,
    )


def make_key(
    name,
    price,
):
    return hashlib.sha1(
        (
            name
            + "|"
            + price
        ).encode(
            "utf-8"
        )
    ).hexdigest()[:16]


def get_outdir():
    downloads = (
        Path.home()
        / "storage"
        / "downloads"
    )

    if downloads.exists():
        outdir = (
            downloads
            / "wechat_mini_export"
        )
    else:
        outdir = (
            Path.home()
            / "wechat_mini_export"
        )

    (
        outdir / "images"
    ).mkdir(
        parents=True,
        exist_ok=True,
    )

    (
        outdir / "debug"
    ).mkdir(
        parents=True,
        exist_ok=True,
    )

    return outdir


def save_csv(
    rows,
    outdir,
):
    path = (
        outdir
        / "products.csv"
    )

    with path.open(
        "w",
        newline="",
        encoding="utf-8-sig",
    ) as f:
        writer = csv.writer(f)

        writer.writerow([
            "商品圖",
            "商品名稱",
            "價格",
        ])

        for row in rows:
            writer.writerow([
                str(
                    row["image"]
                ),
                row["name"],
                row["price"],
            ])

    return path


def save_excel(
    rows,
    outdir,
):
    wb = Workbook()

    ws = wb.active
    ws.title = "商品"

    ws.append([
        "商品圖",
        "商品名稱",
        "價格",
    ])

    ws.column_dimensions[
        "A"
    ].width = 18

    ws.column_dimensions[
        "B"
    ].width = 55

    ws.column_dimensions[
        "C"
    ].width = 16

    for row_no, row in enumerate(
        rows,
        start=2,
    ):
        ws.row_dimensions[
            row_no
        ].height = 95

        ws.cell(
            row_no,
            2,
            row["name"],
        )

        ws.cell(
            row_no,
            3,
            row["price"],
        )

        try:
            img = XLImage(
                str(
                    row["image"]
                )
            )

            img.width = 95
            img.height = 95

            ws.add_image(
                img,
                f"A{row_no}",
            )

        except Exception:
            pass

    path = (
        outdir
        / "products.xlsx"
    )

    wb.save(path)

    return path


def main():
    devices = adb(
        "devices",
        check=False,
    ).stdout

    if "\tdevice" not in devices:
        print(
            "ADB 未連線"
        )
        return

    wm = adb(
        "shell wm size"
    ).stdout

    m = re.search(
        r"(?:Physical size:|Override size:)\s*(\d+)x(\d+)",
        wm,
    )

    if not m:
        print(
            "無法取得螢幕大小"
        )
        return

    sw, sh = map(
        int,
        m.groups(),
    )

    outdir = get_outdir()

    print(
        "螢幕:",
        sw,
        "x",
        sh,
    )

    print(
        "請先保持微信小程序商品列表頁在前景"
    )

    rows = {}
    no_new = 0

    for page in range(
        MAX_SCROLLS
    ):
        # 匯出 UI XML
        adb(
            "shell uiautomator dump "
            "/sdcard/window.xml",
            check=False,
        )

        xml = adb(
            "shell cat "
            "/sdcard/window.xml",
            check=False,
        ).stdout

        if (
            f'package="{WECHAT_PACKAGE}"'
            not in xml
        ):
            print(
                "目前前景可能不是微信"
            )

        # 截圖
        shot = (
            outdir
            / "debug"
            / f"page_{page:04d}.png"
        )

        with shot.open(
            "wb"
        ) as f:
            subprocess.run(
                "adb exec-out screencap -p",
                shell=True,
                stdout=f,
                stderr=subprocess.PIPE,
            )

        screen = Image.open(
            shot
        ).convert(
            "RGB"
        )

        try:
            nodes = parse_nodes(
                xml
            )
        except Exception:
            nodes = []

        products = find_products(
            nodes,
            sw,
        )

        new_count = 0

        for product in products:
            key = make_key(
                product["name"],
                product["price"],
            )

            if key in rows:
                continue

            bounds = (
                get_image_bounds(
                    product,
                    sw,
                    sh,
                )
            )

            x1, y1, x2, y2 = (
                bounds
            )

            if (
                x2 <= x1
                or y2 <= y1
            ):
                continue

            crop = screen.crop(
                bounds
            )

            image_path = (
                outdir
                / "images"
                / f"{key}.jpg"
            )

            crop.save(
                image_path,
                "JPEG",
                quality=92,
            )

            rows[key] = {
                "name":
                    product["name"],

                "price":
                    product["price"],

                "image":
                    image_path,
            }

            new_count += 1

            print(
                "+",
                product["name"],
                product["price"],
            )

        save_csv(
            list(
                rows.values()
            ),
            outdir,
        )

        print(
            "page:",
            page,
            "detected:",
            len(products),
            "new:",
            new_count,
            "total:",
            len(rows),
        )

        if new_count == 0:
            no_new += 1
        else:
            no_new = 0

        if (
            no_new
            >= NO_NEW_LIMIT
        ):
            print(
                "連續多屏沒有新商品，停止"
            )
            break

        # 往上滑
        adb(
            "shell input swipe "
            f"{int(sw * 0.50)} "
            f"{int(sh * 0.78)} "
            f"{int(sw * 0.50)} "
            f"{int(sh * 0.30)} "
            "550",
            check=False,
        )

        time.sleep(
            WAIT_AFTER_SCROLL
        )

    csv_path = save_csv(
        list(
            rows.values()
        ),
        outdir,
    )

    xlsx_path = save_excel(
        list(
            rows.values()
        ),
        outdir,
    )

    print("")
    print("完成")
    print(
        "商品數:",
        len(rows),
    )
    print(
        "Excel:",
        xlsx_path,
    )
    print(
        "CSV:",
        csv_path,
    )
    print(
        "圖片:",
        outdir / "images",
    )


if __name__ == "__main__":
    main()